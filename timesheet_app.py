import streamlit as st
import pandas as pd
import calendar
import holidays
from datetime import date, datetime, time, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="Timesheet Calculator",
    page_icon="🕐",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# CUSTOM CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

:root {
    --bg: #0f1117;
    --surface: #1a1d27;
    --surface2: #242838;
    --accent: #4f9eff;
    --accent2: #ff6b6b;
    --yellow: #ffd166;
    --green: #06d6a0;
    --text: #e8eaf0;
    --text2: #8b8fa8;
    --border: #2e3248;
}

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
    color: var(--text);
}

.stApp { background: var(--bg); }

h1, h2, h3 { font-family: 'IBM Plex Mono', monospace; }

.stSidebar { background: var(--surface) !important; border-right: 1px solid var(--border); }
.stSidebar [data-testid="stSidebarContent"] { padding: 1.5rem 1rem; }

/* Header */
.ts-header {
    background: linear-gradient(135deg, var(--surface) 0%, var(--surface2) 100%);
    border: 1px solid var(--border);
    border-left: 4px solid var(--accent);
    padding: 1.5rem 2rem;
    border-radius: 8px;
    margin-bottom: 1.5rem;
}
.ts-header h1 { color: var(--accent); margin: 0; font-size: 1.6rem; letter-spacing: -0.5px; }
.ts-header p { color: var(--text2); margin: 0.25rem 0 0; font-size: 0.85rem; }

/* Section cards */
.ts-section {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 1.25rem;
    margin-bottom: 1rem;
}
.ts-section-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: var(--accent);
    margin-bottom: 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid var(--border);
}

/* Weekend / holiday cells */
.weekend-cell { background: #3d3800 !important; color: #ffd166 !important; }

/* Metric badges */
.metric-row { display: flex; gap: 1rem; margin: 1rem 0; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 130px;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 0.75rem 1rem;
    text-align: center;
}
.metric-val { font-family: 'IBM Plex Mono', monospace; font-size: 1.6rem; font-weight: 600; }
.metric-lbl { font-size: 0.7rem; color: var(--text2); letter-spacing: 1px; text-transform: uppercase; margin-top: 2px; }

.stButton button {
    background: var(--accent) !important;
    color: #000 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important;
    border: none !important;
    border-radius: 4px !important;
    padding: 0.4rem 1.2rem !important;
}
.stButton button:hover { opacity: 0.85 !important; }

/* Tab style */
.stTabs [data-baseweb="tab-list"] { background: var(--surface) !important; border-radius: 6px; padding: 4px; }
.stTabs [data-baseweb="tab"] { color: var(--text2) !important; }
.stTabs [aria-selected="true"] { color: var(--accent) !important; background: var(--surface2) !important; border-radius: 4px !important; }

/* Input fields */
.stNumberInput input, .stTextInput input, .stSelectbox select {
    background: var(--surface2) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
}

/* Description box */
.desc-box {
    background: var(--surface2);
    border: 1px solid var(--border);
    border-left: 3px solid var(--yellow);
    border-radius: 6px;
    padding: 1rem 1.25rem;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.8rem;
    line-height: 1.7;
    white-space: pre-wrap;
    color: var(--text);
    margin-top: 0.5rem;
}

.tag-inc { color: var(--accent2); font-weight: 600; }
.tag-toil { color: var(--green); font-weight: 600; }
.tag-dl { color: var(--yellow); font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

def get_polish_holidays(year: int, month: int) -> dict:
    """Return dict {date: name} for Polish national holidays in given month."""
    pl_hol = holidays.Poland(years=year)
    return {d: name for d, name in pl_hol.items() if d.month == month}


def is_non_working(d: date, holiday_dates: set) -> bool:
    return d.weekday() >= 5 or d in holiday_dates


def next_working_day(d: date, holiday_dates: set) -> date:
    nd = d + timedelta(days=1)
    while is_non_working(nd, holiday_dates):
        nd += timedelta(days=1)
    return nd


def parse_hhmm(s: str) -> float | None:
    """Parse HH:MM string to float hours."""
    try:
        parts = s.strip().split(":")
        return int(parts[0]) + int(parts[1]) / 60
    except Exception:
        return None


def hours_to_hhmm(h: float) -> str:
    hh = int(h)
    mm = int(round((h - hh) * 60))
    return f"{hh:02d}:{mm:02d}"


# ──────────────────────────────────────────────
# SESSION STATE INIT
# ──────────────────────────────────────────────

def init_state():
    defaults = {
        "incidents": [],
        "overtime_entries": [],
        "break_days": set(),
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_state()

# ──────────────────────────────────────────────
# SIDEBAR — GLOBAL INPUTS
# ──────────────────────────────────────────────

with st.sidebar:
    st.markdown("### 📅 Period")
    col_y, col_m = st.columns(2)
    with col_y:
        year = st.number_input("Year", min_value=2020, max_value=2035, value=date.today().year, step=1)
    with col_m:
        month = st.number_input("Month", min_value=1, max_value=12, value=date.today().month, step=1)

    year, month = int(year), int(month)
    _, days_in_month = calendar.monthrange(year, month)
    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]

    pl_holidays_raw = get_polish_holidays(year, month)
    holiday_dates_set = set(pl_holidays_raw.keys())

    st.markdown("### 🇵🇱 Polish Holidays")
    if pl_holidays_raw:
        hol_override = {}
        for hd, hn in pl_holidays_raw.items():
            active = st.checkbox(f"{hd.strftime('%d')} – {hn}", value=True, key=f"hol_{hd}")
            hol_override[hd] = active
        holiday_dates_set = {d for d, active in hol_override.items() if active}
    else:
        st.caption("No national holidays this month.")

    st.markdown("### ⏱ Daily Hours")
    creative_h = st.number_input("Creative work (h/day)", min_value=0.0, max_value=8.0, value=4.0, step=0.5)
    support_h = st.number_input("Support work (h/day)", min_value=0.0, max_value=8.0, value=4.0, step=0.5)

    st.markdown("### ☕ Break Time")
    add_break = st.checkbox("Add 0.5h break (Others) each working day", value=False)

# ──────────────────────────────────────────────
# MAIN HEADER
# ──────────────────────────────────────────────

st.markdown(f"""
<div class="ts-header">
  <h1>🕐 TIMESHEET CALCULATOR</h1>
  <p>{calendar.month_name[month]} {year} &nbsp;·&nbsp; Poland &nbsp;·&nbsp; {sum(1 for d in all_days if not is_non_working(d, holiday_dates_set))} working days</p>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# TABS
# ──────────────────────────────────────────────

tab1, tab2, tab3, tab4 = st.tabs(["📋 Timesheets", "🚨 Incidents & Overtime", "📝 Description", "📥 Export"])

# ══════════════════════════════════════════════
# TAB 2 — INCIDENTS & OVERTIME (build data first)
# ══════════════════════════════════════════════

with tab2:
    st.markdown('<div class="ts-section-title">ADD INCIDENT / OVERTIME</div>', unsafe_allow_html=True)

    with st.form("entry_form", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            entry_type = st.selectbox("Type", ["Incident", "Overtime"])
        with c2:
            entry_id = st.text_input("ID / Reference", placeholder="INC0001234")
        with c3:
            start_date_inp = st.date_input("Start date", value=date(year, month, 1),
                                           min_value=date(year, month, 1),
                                           max_value=date(year, month, days_in_month))
            start_time_inp = st.text_input("Start time (HH:MM)", value="09:00")
        with c4:
            end_date_inp = st.date_input("End date", value=date(year, month, 1),
                                         min_value=date(year, month, 1),
                                         max_value=date(year, month, days_in_month))
            end_time_inp = st.text_input("End time (HH:MM)", value="10:00")

        submitted = st.form_submit_button("➕ Add Entry")

    if submitted:
        sh = parse_hhmm(start_time_inp)
        eh = parse_hhmm(end_time_inp)
        if sh is None or eh is None:
            st.error("Invalid time format. Use HH:MM")
        else:
            entry = {
                "type": entry_type,
                "id": entry_id or entry_type,
                "start_date": start_date_inp,
                "start_h": sh,
                "end_date": end_date_inp,
                "end_h": eh,
            }
            if entry_type == "Incident":
                st.session_state.incidents.append(entry)
            else:
                st.session_state.overtime_entries.append(entry)
            st.success(f"{entry_type} added!")

    # Display existing entries
    all_entries = (
        [{"Type": e["type"], "ID": e["id"],
          "Start": f"{e['start_date']} {hours_to_hhmm(e['start_h'])}",
          "End": f"{e['end_date']} {hours_to_hhmm(e['end_h'])}",
          "Hours": round(
              (e["end_date"] - e["start_date"]).days * 24 + e["end_h"] - e["start_h"], 2
          )} for e in st.session_state.incidents]
        + [{"Type": e["type"], "ID": e["id"],
            "Start": f"{e['start_date']} {hours_to_hhmm(e['start_h'])}",
            "End": f"{e['end_date']} {hours_to_hhmm(e['end_h'])}",
            "Hours": round(
                (e["end_date"] - e["start_date"]).days * 24 + e["end_h"] - e["start_h"], 2
            )} for e in st.session_state.overtime_entries]
    )

    if all_entries:
        df_entries = pd.DataFrame(all_entries)
        st.dataframe(df_entries, use_container_width=True)
        if st.button("🗑 Clear all entries"):
            st.session_state.incidents = []
            st.session_state.overtime_entries = []
            st.rerun()

# ══════════════════════════════════════════════
# BUSINESS LOGIC — compute per-day data
# ══════════════════════════════════════════════

# Per-day structures
day_data = {}  # date -> dict of computed values

for d in all_days:
    day_data[d] = {
        "is_non_working": is_non_working(d, holiday_dates_set),
        "creative": 0.0,
        "support": 0.0,
        "others": 0.0,      # break or incident-others
        "toil": 0.0,
        "etoil": 0.0,
        "rt": 0.0,
        "ot_support": 0.0,
        "ot_others": 0.0,
        "ot_night": 0.0,
        "dl": False,
        "dl_ref": "",
        "abs_rt": 0.0,
        "abs_toil": 0.0,
        "abs_etoil": 0.0,
        "description_lines": [],
    }

# Fill regular hours for working days
for d in all_days:
    if not day_data[d]["is_non_working"]:
        day_data[d]["creative"] = creative_h
        day_data[d]["support"] = support_h
        if add_break:
            day_data[d]["others"] = 0.5


def process_entry(entry: dict):
    """Apply business logic for a single incident or overtime entry."""
    sd: date = entry["start_date"]
    ed: date = entry["end_date"]
    sh: float = entry["start_h"]
    eh: float = entry["end_h"]
    etype = entry["type"]
    eid = entry["id"]

    # Compute total hours (may span midnight)
    total_h = (ed - sd).days * 24 + eh - sh
    if total_h <= 0:
        return

    # ── INCIDENT RULES ──────────────────────────────────────────────
    if etype == "Incident":
        # "start of working day is 9am; incident at 2am Wednesday = overtime of Tuesday"
        # We re-attribute the incident to the previous calendar day if it starts before 09:00
        # and is during the standard overnight window of the previous working day.

        # Determine the "owner" day for overtime attribution
        if sh < 9.0 and sd.weekday() not in (5, 6):
            # before work start → belongs to previous day
            ot_day = sd - timedelta(days=1)
            # skip back over non-working days
            while ot_day.month == sd.month and is_non_working(ot_day, holiday_dates_set):
                ot_day -= timedelta(days=1)
        else:
            ot_day = sd  # same day after 17:00

        if ot_day not in day_data:
            ot_day = sd  # fallback

        # Is the incident during a weekend or holiday? → DL
        # weekend = 00:00 Sat – 24:00 Sun, or on a public holiday day
        is_dl_eligible = (
            (sd.weekday() >= 5)
            or (sd in holiday_dates_set)
            or (ed.weekday() >= 5 and eh > 0)
            or (ed in holiday_dates_set and eh > 0)
        )

        # Night hours between 22:00 and 04:00
        def night_hours_in_span(s_d, s_h, e_d, e_h):
            nh = 0.0
            current = datetime(s_d.year, s_d.month, s_d.day) + timedelta(hours=s_h)
            end_dt = datetime(e_d.year, e_d.month, e_d.day) + timedelta(hours=e_h)
            while current < end_dt:
                nxt = min(current + timedelta(hours=0.5), end_dt)
                mid = current + timedelta(minutes=15)
                h_mid = mid.hour + mid.minute / 60
                if h_mid >= 22 or h_mid < 4:
                    nh += (nxt - current).total_seconds() / 3600
                current = nxt
            return nh

        night_h = night_hours_in_span(sd, sh, ed, eh)

        # Record overtime on ot_day
        if ot_day in day_data:
            day_data[ot_day]["ot_others"] += total_h
            day_data[ot_day]["ot_night"] += night_h

        # TOIL on next working day after ot_day
        toil_day = next_working_day(ot_day, holiday_dates_set)
        if toil_day in day_data:
            day_data[toil_day]["abs_toil"] += total_h
            day_data[toil_day]["abs_etoil"] += round(total_h * 0.5, 2)
            if night_h > 0:
                day_data[toil_day]["abs_rt"] += night_h

        # DL
        if is_dl_eligible:
            # Find the next working day for DL
            dl_day = next_working_day(ed, holiday_dates_set)
            if dl_day in day_data:
                day_data[dl_day]["dl"] = True
                day_data[dl_day]["dl_ref"] = eid

        # Description lines
        reported_on = ""
        if ot_day != sd:
            reported_on = f" (reported on {ot_day.strftime('%d/%m')})"
        line = (
            f"{sd.strftime('%d/%m/%Y')} - INC - "
            f"{hours_to_hhmm(sh)} to {hours_to_hhmm(eh)}{reported_on} - "
            f"{hours_to_hhmm(total_h)}"
        )
        if ot_day in day_data:
            day_data[ot_day]["description_lines"].append(("INC", line))

        toil_line = (
            f"{toil_day.strftime('%d/%m/%Y')} - "
            f"TOIL {hours_to_hhmm(total_h)} + E-TOIL {hours_to_hhmm(round(total_h*0.5,2))}"
        )
        if night_h > 0:
            toil_line += f" + RT {hours_to_hhmm(night_h)}"
        if toil_day in day_data:
            day_data[toil_day]["description_lines"].append(("TOIL", toil_line))

        if is_dl_eligible and dl_day in day_data:
            day_data[dl_day]["description_lines"].append(
                ("DL", f"{dl_day.strftime('%d/%m/%Y')} - DL FOR INC ON {sd.strftime('%d/%m/%Y')}")
            )

    # ── OVERTIME RULES ───────────────────────────────────────────────
    else:
        # Overtime does NOT give TOIL
        if sd in day_data:
            # Only log overtime outside 09-17 window
            day_data[sd]["ot_support"] += total_h
            day_data[sd]["description_lines"].append(
                ("OT", f"{sd.strftime('%d/%m/%Y')} - OT - "
                       f"{hours_to_hhmm(sh)} to {hours_to_hhmm(eh)} - "
                       f"{hours_to_hhmm(total_h)}")
            )


for inc in st.session_state.incidents:
    process_entry(inc)

for ot in st.session_state.overtime_entries:
    process_entry(ot)

# DL absent hours
for d, dd in day_data.items():
    if dd["dl"] and not dd["is_non_working"]:
        dd["abs_toil"] += 8.0  # DL = 8h absence

# ══════════════════════════════════════════════
# TAB 1 — TIMESHEETS (tables)
# ══════════════════════════════════════════════

with tab1:

    # ── TABLE 1: Regular Hours ────────────────────────────────────────
    st.markdown('<div class="ts-section-title">TABLE 1 — REGULAR WORKING HOURS (ECoE / Project)</div>', unsafe_allow_html=True)

    rows_t1 = {
        "ECoE / Creative Work": [],
        "ECoE / Support": [],
        "ECoE / Others": [],
        "Absence / RT": [],
        "Absence / TOIL": [],
        "Absence / E-TOIL": [],
        "── TOTAL ──": [],
    }

    col_labels = []
    for d in all_days:
        wd = calendar.day_abbr[d.weekday()]
        col_labels.append(f"{d.day}\n{wd}")

    for d in all_days:
        dd = day_data[d]
        nw = dd["is_non_working"]
        rows_t1["ECoE / Creative Work"].append("" if nw else dd["creative"])
        rows_t1["ECoE / Support"].append("" if nw else dd["support"])
        rows_t1["ECoE / Others"].append("" if nw else (dd["others"] if dd["others"] else ""))
        rows_t1["Absence / RT"].append("" if nw else (dd["abs_rt"] if dd["abs_rt"] else ""))
        rows_t1["Absence / TOIL"].append("" if nw else (dd["abs_toil"] if dd["abs_toil"] else ""))
        rows_t1["Absence / E-TOIL"].append("" if nw else (dd["abs_etoil"] if dd["abs_etoil"] else ""))

        if nw:
            rows_t1["── TOTAL ──"].append("")
        else:
            total = sum(filter(None, [
                dd["creative"], dd["support"], dd["others"],
                dd["abs_rt"], dd["abs_toil"], dd["abs_etoil"]
            ]))
            rows_t1["── TOTAL ──"].append(total if total else "")

    df1 = pd.DataFrame(rows_t1, index=col_labels).T

    # Style non-working days
    def style_t1(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for i, d in enumerate(all_days):
            col = col_labels[i]
            if is_non_working(d, holiday_dates_set):
                styles[col] = "background-color: #3d3800; color: #ffd166;"
            elif col in df.columns:
                if df.index.get_loc("── TOTAL ──") is not None:
                    pass
        # Bold total row
        styles.loc["── TOTAL ──"] = styles.loc["── TOTAL ──"].apply(
            lambda x: x + " font-weight: bold; border-top: 2px solid #4f9eff;"
        )
        return styles

    st.dataframe(
        df1.style.apply(lambda _: [
            "background-color: #3d3800; color: #ffd166;"
            if is_non_working(all_days[col_labels.index(c)], holiday_dates_set) else ""
            for c in df1.columns
        ], axis=None).apply(
            lambda row: ["font-weight: bold; border-top: 2px solid #4f9eff;" if row.name == "── TOTAL ──" else "" for _ in row],
            axis=1
        ),
        use_container_width=True,
        height=280,
    )

    # Summary T1
    total_project = sum(
        day_data[d]["creative"] + day_data[d]["support"] + day_data[d]["others"]
        for d in all_days if not day_data[d]["is_non_working"]
    )
    total_absence = sum(
        day_data[d]["abs_rt"] + day_data[d]["abs_toil"] + day_data[d]["abs_etoil"]
        for d in all_days
    )

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card">
            <div class="metric-val" style="color:#4f9eff">{total_project:.1f}h</div>
            <div class="metric-lbl">Project Hours</div>
        </div>
        <div class="metric-card">
            <div class="metric-val" style="color:#ff6b6b">{total_absence:.1f}h</div>
            <div class="metric-lbl">Absence Hours</div>
        </div>
        <div class="metric-card">
            <div class="metric-val" style="color:#06d6a0">{total_project + total_absence:.1f}h</div>
            <div class="metric-lbl">Total</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # ── TABLE 2: Overtime Hours ────────────────────────────────────────
    st.markdown('<div class="ts-section-title">TABLE 2 — OVERTIME HOURS</div>', unsafe_allow_html=True)

    rows_t2 = {
        "ECoE / Support": [],
        "ECoE / Others": [],
        "  └ Night Time": [],
        "TOIL (from Table 1)": [],
        "DL (Day in Lieu)": [],
    }

    for d in all_days:
        dd = day_data[d]
        rows_t2["ECoE / Support"].append(dd["ot_support"] if dd["ot_support"] else "")
        rows_t2["ECoE / Others"].append(dd["ot_others"] if dd["ot_others"] else "")
        rows_t2["  └ Night Time"].append(dd["ot_night"] if dd["ot_night"] else "")
        rows_t2["TOIL (from Table 1)"].append(dd["abs_toil"] if dd["abs_toil"] else "")
        rows_t2["DL (Day in Lieu)"].append("YES" if dd["dl"] else "")

    df2 = pd.DataFrame(rows_t2, index=col_labels).T

    def highlight_ot(df):
        def col_style(col):
            d = all_days[col_labels.index(col)]
            if is_non_working(d, holiday_dates_set):
                return ["background-color: #3d3800; color: #ffd166;"] * len(df)
            return [""] * len(df)
        return df.apply(col_style, axis=0)

    st.dataframe(
        df2.style.apply(highlight_ot),
        use_container_width=True,
        height=220,
    )

    total_ot = sum(
        day_data[d]["ot_support"] + day_data[d]["ot_others"]
        for d in all_days
    )
    total_toil = sum(day_data[d]["abs_toil"] for d in all_days)
    total_dl = sum(1 for d in all_days if day_data[d]["dl"])

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card">
            <div class="metric-val" style="color:#ff6b6b">{total_ot:.1f}h</div>
            <div class="metric-lbl">Total Overtime</div>
        </div>
        <div class="metric-card">
            <div class="metric-val" style="color:#06d6a0">{total_toil:.1f}h</div>
            <div class="metric-lbl">Total TOIL</div>
        </div>
        <div class="metric-card">
            <div class="metric-val" style="color:#ffd166">{total_dl}</div>
            <div class="metric-lbl">DL Days</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════
# TAB 3 — DESCRIPTION
# ══════════════════════════════════════════════

with tab3:
    st.markdown('<div class="ts-section-title">AUTO-GENERATED DESCRIPTION</div>', unsafe_allow_html=True)

    desc_lines = []
    for d in sorted(all_days):
        for tag, line in day_data[d]["description_lines"]:
            desc_lines.append(line)

    desc_lines.append("SPREADSHEET IS ATTACHED")

    desc_text = "\n".join(desc_lines)

    st.markdown(f'<div class="desc-box">{desc_text}</div>', unsafe_allow_html=True)

    st.text_area(
        "Edit Description (optional)",
        value=desc_text,
        height=300,
        key="desc_edit",
        label_visibility="collapsed",
    )

# ══════════════════════════════════════════════
# TAB 4 — EXPORT
# ══════════════════════════════════════════════

def build_excel() -> BytesIO:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Regular Hours"
    ws2 = wb.create_sheet("Overtime Hours")
    ws3 = wb.create_sheet("Description")

    YELLOW_FILL = PatternFill("solid", fgColor="FFD166")
    HEADER_FILL = PatternFill("solid", fgColor="1A1D27")
    TOTAL_FILL = PatternFill("solid", fgColor="242838")
    HEADER_FONT = Font(name="Calibri", bold=True, color="4F9EFF", size=9)
    TOTAL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    BODY_FONT = Font(name="Calibri", size=9, color="E8EAF0")
    YELLOW_FONT = Font(name="Calibri", bold=True, color="3D3800", size=9)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="2E3248")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_table(ws, rows_dict, all_days, col_labels, holiday_dates_set):
        # Header row
        ws.cell(1, 1, "Category").font = HEADER_FONT
        ws.cell(1, 1).fill = HEADER_FILL
        ws.cell(1, 1).alignment = CENTER
        ws.cell(1, 1).border = border
        ws.column_dimensions["A"].width = 22

        for ci, (d, lbl) in enumerate(zip(all_days, col_labels), start=2):
            c = ws.cell(1, ci, lbl)
            c.alignment = CENTER
            c.border = border
            nw = is_non_working(d, holiday_dates_set)
            if nw:
                c.fill = YELLOW_FILL
                c.font = YELLOW_FONT
            else:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
            ws.column_dimensions[get_column_letter(ci)].width = 5.5

        for ri, (row_name, values) in enumerate(rows_dict.items(), start=2):
            ws.cell(ri, 1, row_name).font = HEADER_FONT
            ws.cell(ri, 1).fill = TOTAL_FILL if "TOTAL" in row_name else HEADER_FILL
            ws.cell(ri, 1).alignment = CENTER
            ws.cell(ri, 1).border = border

            for ci, (val, d) in enumerate(zip(values, all_days), start=2):
                c = ws.cell(ri, ci, val if val != "" else None)
                nw = is_non_working(d, holiday_dates_set)
                if nw:
                    c.fill = YELLOW_FILL
                    c.font = YELLOW_FONT
                elif "TOTAL" in row_name:
                    c.fill = TOTAL_FILL
                    c.font = TOTAL_FONT
                else:
                    c.fill = PatternFill("solid", fgColor="0F1117")
                    c.font = BODY_FONT
                c.alignment = CENTER
                c.border = border

    write_table(ws1, rows_t1, all_days, col_labels, holiday_dates_set)
    write_table(ws2, rows_t2, all_days, col_labels, holiday_dates_set)

    # Description sheet
    ws3.column_dimensions["A"].width = 80
    for i, line in enumerate(desc_lines, start=1):
        c = ws3.cell(i, 1, line)
        c.font = Font(name="Courier New", size=10, color="E8EAF0")
        c.fill = PatternFill("solid", fgColor="0F1117")
        ws3.row_dimensions[i].height = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


with tab4:
    st.markdown('<div class="ts-section-title">EXPORT TIMESHEET</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#1a1d27; border:1px solid #2e3248; border-radius:6px; padding:1rem 1.25rem; margin-bottom:1rem;">
        <p style="margin:0; color:#8b8fa8; font-size:0.85rem;">
        Export generates an <strong style="color:#4f9eff">.xlsx spreadsheet</strong> with:<br>
        • Table 1 — Regular Working Hours<br>
        • Table 2 — Overtime Hours<br>
        • Description sheet (auto-generated text)<br>
        Weekend & holiday columns are highlighted in yellow.
        </p>
    </div>
    """, unsafe_allow_html=True)

    excel_buf = build_excel()
    fname = f"timesheet_{year}_{month:02d}.xlsx"

    st.download_button(
        label="📥 Download Excel Timesheet",
        data=excel_buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.caption(f"File: `{fname}`")
